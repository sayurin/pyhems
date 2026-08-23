#!/usr/bin/env -S uv run
# /// script
# requires-python = ">=3.13"
# dependencies = [
#   "openpyxl",
#   "pyyaml",
# ]
# ///
# ruff: noqa: T201
"""Generate ECHONET Lite definitions from MRA data.

This script:
1. Reads MRA data from the local mra/ directory
2. Parses MRA JSON to extract device and property specifications
3. Loads custom definitions from custom_definitions.yaml
4. Generates _definitions_generated.py for runtime entity creation

Run with: uv run scripts/generate_definitions.py

Output file:
- src/pyhems/_definitions_generated.py
"""

from __future__ import annotations

import dataclasses
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, TypeAlias

import yaml
from openpyxl import load_workbook

# ============================================================================
# Constants
# ============================================================================

PYHEMS_DIR = Path(__file__).parent.parent / "src" / "pyhems"
MRA_DIR = Path(__file__).parent.parent / "mra"
CUSTOM_DEFINITIONS_FILE = Path(__file__).parent / "custom_definitions.yaml"
MANUFACTURER_CODES_FILE = Path(__file__).parent / "manufacturer_codes.yaml"
PROPERTY_ROLES_FILE = Path(__file__).parent / "property_roles.xlsx"

# Maps known two-value state enum keys to the generated boolean keys. The
# input order is deliberately irrelevant: boolean consumers must only use the
# normalized true/false keys.
_BOOLEAN_ENUM_KEY_NORMALIZATIONS: dict[frozenset[str], dict[str, str]] = {
    frozenset({"anyOpen", "allClose"}): {"anyOpen": "true", "allClose": "false"},
    frozenset({"enable", "disable"}): {"enable": "true", "disable": "false"},
    frozenset({"enabled", "disabled"}): {"enabled": "true", "disabled": "false"},
    frozenset({"generationOn", "generationOff"}): {
        "generationOn": "true",
        "generationOff": "false",
    },
    frozenset({"necessary", "notNecessary"}): {
        "necessary": "true",
        "notNecessary": "false",
    },
    frozenset({"normal", "warning"}): {"normal": "false", "warning": "true"},
    frozenset({"on", "off"}): {"on": "true", "off": "false"},
    frozenset({"open", "close"}): {"open": "true", "close": "false"},
    frozenset({"open", "closed"}): {"open": "true", "closed": "false"},
    frozenset({"permitted", "prohibited"}): {
        "permitted": "true",
        "prohibited": "false",
    },
    frozenset({"ready", "busy"}): {"ready": "false", "busy": "true"},
    frozenset({"running", "stopped"}): {"running": "true", "stopped": "false"},
    frozenset({"running", "suspension"}): {
        "running": "true",
        "suspension": "false",
    },
}

# Known two-value state enums that represent mutually exclusive modes rather
# than a boolean state. These must retain their source keys and use EnumCodec.
_NON_BOOLEAN_TWO_VALUE_ENUM_KEYS = frozenset(
    {
        frozenset({"airConditioning", "blowing"}),
        frozenset({"cooling", "heating"}),
        frozenset({"cooling", "nonCooling"}),
        frozenset({"devicePoint", "powerReceivingPoint"}),
        frozenset({"loadFollowing", "maximumRating"}),
        frozenset({"builtIn", "separate"}),
        frozenset({"freezing", "refrigeration"}),
    }
)

# ============================================================================
# Load definitions dataclasses from pyhems/definitions.py
#
# Loaded via sys.path so definitions.py can be imported standalone, without
# pulling in the full pyhems package and its runtime dependencies
# (codecs, device_manager, etc.).
# ============================================================================

sys.path.insert(0, str(PYHEMS_DIR))
try:
    import definitions as _defs_mod
finally:
    sys.path.pop(0)

EnumValue: TypeAlias = _defs_mod.EnumValue  # noqa: UP040
EntityDefinition: TypeAlias = _defs_mod.EntityDefinition  # noqa: UP040
DeviceDefinition: TypeAlias = _defs_mod.DeviceDefinition  # noqa: UP040
ManufacturerDefinition: TypeAlias = _defs_mod.ManufacturerDefinition  # noqa: UP040
DefinitionsRegistry: TypeAlias = _defs_mod.DefinitionsRegistry  # noqa: UP040
NumericValueEntry: TypeAlias = _defs_mod.NumericValueEntry  # noqa: UP040
PropertyRole: TypeAlias = _defs_mod.PropertyRole  # noqa: UP040
ScalarDefinition: TypeAlias = _defs_mod.ScalarDefinition  # noqa: UP040
ObjectField: TypeAlias = _defs_mod.ObjectField  # noqa: UP040
ObjectDefinition: TypeAlias = _defs_mod.ObjectDefinition  # noqa: UP040
ArrayDefinition: TypeAlias = _defs_mod.ArrayDefinition  # noqa: UP040
OneOfDefinition: TypeAlias = _defs_mod.OneOfDefinition  # noqa: UP040
PropertyValueDefinition: TypeAlias = _defs_mod.PropertyValueDefinition  # noqa: UP040
CollectionBinding: TypeAlias = _defs_mod.CollectionBinding  # noqa: UP040
CollectionIndex: TypeAlias = _defs_mod.CollectionIndex  # noqa: UP040


# ============================================================================
# Internal Build Containers
# ============================================================================


@dataclass
class _DeviceBuild:
    """Mutable container for building device-specific entity lists."""

    name_en: str
    name_ja: str
    entities: list[EntityDefinition]  # device-specific only (excludes common)
    structured_values: dict[int, PropertyValueDefinition] = dataclasses.field(
        default_factory=dict
    )
    collection_bindings: list[CollectionBinding] = dataclasses.field(
        default_factory=list
    )


@dataclass
class _DefinitionsBuild:
    """Mutable top-level container for the full definitions build."""

    version: str
    mra_version: str
    common: list[EntityDefinition]
    devices: dict[int, _DeviceBuild]
    manufacturers: dict[int, ManufacturerDefinition]
    mra_epcs: dict[int, frozenset[int]]
    common_epcs: frozenset[int]


# ============================================================================
# Internal MRA Parsing Model
# ============================================================================


@dataclass
class _MRAProperty:
    """Parsed MRA property data (internal, not exported)."""

    epc: int
    name_en: str
    name_ja: str
    description_en: str | None
    description_ja: str | None
    get: str
    set: str
    data_type: str | None
    mra_format: str | None  # e.g., "uint16", "int8"
    mra_unit: str | None  # e.g., "W", "Wh", "Celsius"
    mra_minimum: float | None
    mra_maximum: float | None
    mra_multiple_of: float | None  # scale factor, e.g., 0.1 for tenths
    enum_values: list[EnumValue]
    has_level_enums: bool = False  # True if any level type was processed
    numeric_values: tuple[NumericValueEntry, ...] = ()
    coefficient_epcs: tuple[int, ...] = ()


# ============================================================================
# Utility Functions
# ============================================================================


def _normalize_trailing_number(name: str) -> str:
    """Insert a space before trailing digits if missing.

    e.g. 'Lock setting1' -> 'Lock setting 1'
    Only inserts when at least two letters immediately precede the trailing digits,
    avoiding short codes like 'n1'.
    """
    return re.sub(r"(?<=[a-zA-Z]{2})(\d+)$", r" \1", name)


def _normalize_two_value_enum_keys(
    enum_values: tuple[EnumValue, ...],
    class_code: int,
    epc: int,
    manufacturer_code: int | None = None,
) -> tuple[EnumValue, ...]:
    """Normalize known boolean enum keys and reject unclassified two-value enums."""
    if len(enum_values) != 2:
        return enum_values

    keys = frozenset(value.key for value in enum_values)
    if keys == {"true", "false"} or keys in _NON_BOOLEAN_TWO_VALUE_ENUM_KEYS:
        return enum_values

    if normalization := _BOOLEAN_ENUM_KEY_NORMALIZATIONS.get(keys):
        return tuple(
            dataclasses.replace(value, key=normalization[value.key])
            for value in enum_values
        )

    manufacturer = (
        f", manufacturer 0x{manufacturer_code:06X}"
        if manufacturer_code is not None
        else ""
    )
    raise ValueError(
        f"Unclassified two-value enum for class 0x{class_code:04X} EPC 0x{epc:02X}"
        f"{manufacturer}: {sorted(keys)!r}. Classify it as boolean or non-boolean."
    )


def _parse_hex_int(value: Any) -> int | None:
    """Parse a hex string or int to integer.

    Args:
        value: String like "0x0135" or integer.

    Returns:
        Integer value or None if parsing fails.
    """
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        try:
            return int(value, 0)  # auto-detect base (0x for hex)
        except ValueError:
            return None
    return None


# ============================================================================
# MRA Parsing Functions
# ============================================================================


def _parse_mra_property(
    prop_data: dict[str, Any], definitions: dict[str, Any]
) -> _MRAProperty:
    """Parse a single MRA property definition."""
    # MRA properties always have these required keys: epc, propertyName, accessRule, data
    epc = int(prop_data["epc"], 16)

    name_data = prop_data["propertyName"]
    name_en = _normalize_trailing_number(name_data["en"])
    name_ja = name_data["ja"]
    descriptions_data = prop_data["descriptions"]
    description_en = descriptions_data["en"]
    description_ja = descriptions_data["ja"]

    # Parse access rules
    access = prop_data["accessRule"]
    get_val = access["get"]
    set_val = access["set"]

    # Parse data specification
    data_spec = prop_data["data"]

    # Collect all data specs to process (for oneOf, process all elements)
    data_specs_to_process: list[dict[str, Any]] = []
    if "oneOf" in data_spec:
        one_of = data_spec["oneOf"]
        if one_of and isinstance(one_of, list):
            data_specs_to_process.extend(one_of)
    else:
        data_specs_to_process.append(data_spec)

    # Resolve $ref (merging sibling keys such as "coefficient" that MRA places
    # alongside "$ref" rather than inside the referenced definition) and
    # collect all resolved specs
    resolved_specs: list[dict[str, Any]] = []
    for spec in data_specs_to_process:
        if "$ref" in spec:
            ref = spec["$ref"]
            if ref.startswith("#/definitions/"):
                def_name = ref.replace("#/definitions/", "")
                if def_name in definitions:
                    resolved_specs.append(
                        {
                            **definitions[def_name],
                            **{k: v for k, v in spec.items() if k != "$ref"},
                        }
                    )
        else:
            resolved_specs.append(spec)

    # Use first spec for primary type inference
    assert resolved_specs, f"No resolved specs for EPC 0x{epc:02X}"
    data_spec = resolved_specs[0]
    data_type: str | None = data_spec["type"]

    # Extract numeric type info from MRA definition
    mra_format: str | None = None
    mra_unit: str | None = None
    mra_minimum: float | None = None
    mra_maximum: float | None = None
    mra_multiple_of: float | None = None
    coefficient_epcs: tuple[int, ...] = ()

    if data_type == "number":
        mra_format = data_spec.get("format")
        mra_unit = data_spec.get("unit")
        mra_minimum = data_spec.get("minimum")
        mra_maximum = data_spec.get("maximum")
        mra_multiple_of = data_spec.get("multiple") or data_spec.get("multipleOf")
        coefficient_epcs = tuple(int(c, 16) for c in data_spec.get("coefficient", []))

    # MRA "numericValue" type: edt byte -> float multiplier table (e.g. EPC
    # 0xC2 "Unit for cumulative amounts of electric energy"). Referenced by
    # other properties via their "coefficient" array.
    numeric_values: tuple[NumericValueEntry, ...] = ()
    if data_type == "numericValue":
        numeric_values = tuple(
            NumericValueEntry(edt=int(item["edt"], 16), value=item["numericValue"])
            for item in data_spec.get("enum", [])
        )

    # Parse enum values from all specs
    enum_values: list[EnumValue] = []
    has_level_enums = False

    for spec in resolved_specs:
        spec_type = spec.get("type")

        # Handle level type
        if spec_type == "level":
            base = int(spec["base"], 16)
            maximum = spec["maximum"]
            if base >= 256 or maximum > 32:
                continue
            for i in range(maximum):
                edt = base + i
                level_num = i + 1
                enum_key = f"level_{level_num}"
                enum_values.append(
                    EnumValue(
                        edt=edt,
                        key=enum_key,
                        name_en=f"Level {level_num}",
                        name_ja=f"レベル{level_num}",
                    )
                )
            data_type = "state"
            has_level_enums = True

        # Handle state type with enum
        elif spec_type == "state":
            for item in spec.get("enum", []):
                edt_str = item["edt"]
                val_name = item["name"]
                if "..." in edt_str:
                    continue
                edt = int(edt_str, 16)
                descriptions = item["descriptions"]
                enum_values.append(
                    EnumValue(
                        edt=edt,
                        key=val_name,
                        name_en=_normalize_trailing_number(descriptions["en"]),
                        name_ja=descriptions["ja"],
                    )
                )

    return _MRAProperty(
        epc=epc,
        name_en=name_en,
        name_ja=name_ja,
        description_en=description_en,
        description_ja=description_ja,
        get=get_val,
        set=set_val,
        data_type=data_type,
        mra_format=mra_format,
        mra_unit=mra_unit,
        mra_minimum=mra_minimum,
        mra_maximum=mra_maximum,
        mra_multiple_of=mra_multiple_of,
        enum_values=enum_values,
        has_level_enums=has_level_enums,
        numeric_values=numeric_values,
        coefficient_epcs=coefficient_epcs,
    )


# ============================================================================
# Entity Building Functions
# ============================================================================


def _build_entity_from_property(
    class_code: int,
    prop: _MRAProperty,
) -> EntityDefinition | None:
    """Build EntityDefinition directly from an MRA property."""
    is_readable = prop.get in ("required", "required_c", "required_o", "optional")
    is_writable = prop.set in ("required", "required_c", "required_o", "optional")
    assert is_readable or is_writable, (
        f"Property for class 0x{class_code:04X} EPC 0x{prop.epc:02X} "
        "is neither readable nor writable"
    )

    # A "number" property is sensor-eligible regardless of whether MRA gives
    # it a unit: unitless counts/indices (e.g. class 0x0287 EPC 0xB1/0xB8
    # channel counts) are just as meaningful as unit-bearing measurements.
    # This matches _build_entities_from_object_property(), which never
    # required a unit for its fixed-layout object fields either.
    is_sensor = prop.data_type == "number"
    is_state = prop.data_type == "state"
    is_numeric_value = prop.data_type == "numericValue"

    if not is_sensor and not is_state and not is_numeric_value:
        return None

    # Filter out level-based properties with too many enum values (>16)
    # Pure state enums (e.g., washing machine courses) are kept regardless of count
    if is_state and prop.has_level_enums and len(prop.enum_values) > 16:
        return None

    # Skip entities where different EDT bytes share the same key name.
    # Such enums make encode() non-deterministic and produce an unusable codec.
    if is_state:
        keys = [ev.key for ev in prop.enum_values]
        if len(keys) != len(set(keys)):
            return None

    name_en = prop.name_en
    assert name_en, (
        f"Missing English name for class 0x{class_code:04X} EPC 0x{prop.epc:02X}"
    )

    enum_vals: list[EnumValue] = prop.enum_values
    assert not is_state or enum_vals, (
        f"state entity for class 0x{class_code:04X} EPC 0x{prop.epc:02X} "
        f"({name_en}) has no enum_values"
    )

    # For sensors: filter out out-of-range enum_values (special markers like
    # "Unmeasurable", "Not measured" that sit outside the numeric range)
    if (
        enum_vals
        and is_sensor
        and prop.mra_minimum is not None
        and prop.mra_maximum is not None
    ):
        enum_vals = [
            ev for ev in enum_vals if prop.mra_minimum <= ev.edt <= prop.mra_maximum
        ]

    return EntityDefinition(
        id=f"class_{class_code:04x}_epc_{prop.epc:02x}",
        epc=prop.epc,
        name_en=name_en,
        name_ja=prop.name_ja,
        get=prop.get,
        set=prop.set,
        description_en=prop.description_en,
        description_ja=prop.description_ja,
        format=prop.mra_format if is_sensor else None,
        unit=prop.mra_unit if is_sensor else None,
        minimum=prop.mra_minimum if is_sensor else None,
        maximum=prop.mra_maximum if is_sensor else None,
        multiple_of=(
            prop.mra_multiple_of
            if is_sensor and prop.mra_multiple_of is not None
            else 1.0
        ),
        enum_values=_normalize_two_value_enum_keys(
            tuple(enum_vals), class_code, prop.epc
        ),
        numeric_values=prop.numeric_values if is_numeric_value else None,
        coefficient_epcs=(prop.coefficient_epcs or None) if is_sensor else None,
    )


# ============================================================================
# Fixed-layout "object" property splitting
#
# Some MRA properties pack several scalar numeric fields into a single EDT
# (e.g. class 0x0287 EPC 0xD0 "Measurement channel 1": cumulative kWh + two
# instantaneous currents). This mirrors the byte_offset pattern already used
# for manufacturer-specific entries in custom_definitions.yaml, but derives it
# directly from the MRA "object" type instead of hand-written YAML.
#
# Variable-length shapes (MRA "array" fields, e.g. the chunked channel lists
# on EPC 0xB3/0xB5/0xC3/0xC4) are intentionally unsupported here and cause
# _build_entities_from_object_property() to return None, leaving the
# property dropped exactly as before this feature was added.
# ============================================================================

_OBJECT_FIELD_BYTE_SIZE: dict[str, int] = {
    "uint8": 1,
    "int8": 1,
    "uint16": 2,
    "int16": 2,
    "uint32": 4,
    "int32": 4,
}

# ECHONET Lite standard time-related types (byte width per the ECHONET Lite
# specification): "date" = year(2)+month(1)+day(1), "time" = relative
# hour(1)+minute(1)+second(1), "date-time" = date(4)+time(3).
_RAW_TIME_TYPE_SIZE: dict[str, int] = {
    "date": 4,
    "time": 3,
    "date-time": 7,
}


def _resolve_ref(spec: dict[str, Any], definitions: dict[str, Any]) -> dict[str, Any]:
    """Resolve a single-level MRA '$ref' pointer against shared definitions.

    Sibling keys placed alongside "$ref" (e.g. "coefficient", "overflowCode")
    are merged into the resolved result rather than discarded, matching the
    MRA convention of annotating a shared definition per use site.
    """
    if "$ref" in spec:
        ref = spec["$ref"]
        if ref.startswith("#/definitions/"):
            resolved = definitions[ref.replace("#/definitions/", "")]
            return {**resolved, **{k: v for k, v in spec.items() if k != "$ref"}}
    return spec


# ============================================================================
# Recursive structured value-tree parsing
#
# Builds a PropertyValueDefinition tree (see pyhems.definitions) for any MRA
# "data"/"element" spec, including variable-length "array" fields that
# _build_entities_from_object_property() above cannot flatten into scalar
# EntityDefinitions. Used to preserve properties that would otherwise be
# dropped entirely (e.g. class 0x0287 EPC 0xB3/0xB7/0xBA/0xBE channel lists).
# ============================================================================


def _parse_state_enum_values(spec: dict[str, Any]) -> list[EnumValue]:
    """Parse a resolved MRA 'state' spec's enum list into EnumValue entries."""
    enum_values: list[EnumValue] = []
    for item in spec.get("enum", []):
        edt_str = item["edt"]
        if "..." in edt_str:
            continue
        descriptions = item["descriptions"]
        enum_values.append(
            EnumValue(
                edt=int(edt_str, 16),
                key=item["name"],
                name_en=_normalize_trailing_number(descriptions["en"]),
                name_ja=descriptions["ja"],
            )
        )
    return enum_values


def _parse_value_spec(
    spec: dict[str, Any], definitions: dict[str, Any]
) -> PropertyValueDefinition:
    """Recursively parse an MRA data/element spec into a value definition tree."""
    if "oneOf" in spec:
        return OneOfDefinition(
            options=tuple(
                _parse_value_spec(option, definitions) for option in spec["oneOf"]
            )
        )

    resolved = _resolve_ref(spec, definitions)
    data_type = resolved.get("type")

    if data_type == "array":
        return ArrayDefinition(
            item=_parse_value_spec(resolved["items"], definitions),
            item_size=resolved["itemSize"],
            min_items=resolved.get("minItems"),
            max_items=resolved.get("maxItems"),
        )

    if data_type == "object":
        return ObjectDefinition(
            fields=tuple(
                ObjectField(
                    key=field["shortName"],
                    name_en=_normalize_trailing_number(field["elementName"]["en"]),
                    name_ja=field["elementName"]["ja"],
                    value=_parse_value_spec(field["element"], definitions),
                )
                for field in resolved["properties"]
            )
        )

    if data_type == "numericValue":
        return ScalarDefinition(
            size=resolved["size"],
            numeric_values=tuple(
                NumericValueEntry(edt=int(item["edt"], 16), value=item["numericValue"])
                for item in resolved.get("enum", [])
            ),
        )

    if data_type == "state":
        return ScalarDefinition(
            size=resolved["size"], enum_values=tuple(_parse_state_enum_values(resolved))
        )

    if data_type == "level":
        base = int(resolved["base"], 16)
        maximum = resolved["maximum"]
        enum_values = tuple(
            EnumValue(
                edt=base + i,
                key=f"level_{i + 1}",
                name_en=f"Level {i + 1}",
                name_ja=f"レベル{i + 1}",
            )
            for i in range(maximum)
            if base + i < 256
        )
        return ScalarDefinition(size=1, enum_values=enum_values)

    if data_type == "number":
        mra_format: str | None = resolved.get("format")
        return ScalarDefinition(
            size=_OBJECT_FIELD_BYTE_SIZE.get(mra_format, 1) if mra_format else 1,
            format=mra_format,
            unit=resolved.get("unit"),
            minimum=resolved.get("minimum"),
            maximum=resolved.get("maximum"),
            multiple_of=resolved.get("multiple") or resolved.get("multipleOf") or 1.0,
            coefficient_epcs=tuple(int(c, 16) for c in resolved.get("coefficient", [])),
        )

    # ECHONET Lite standard time-related types: fixed byte widths per the
    # ECHONET Lite specification, not documented via a "size" key in MRA
    # (unlike "state"/"numericValue"). Not decoded numerically here (no
    # format), just preserved with the correct width so sibling oneOf
    # sentinel states (e.g. state_Unknown_FFFFFFFF alongside "date") agree
    # on byte size.
    if data_type in _RAW_TIME_TYPE_SIZE:
        return ScalarDefinition(size=_RAW_TIME_TYPE_SIZE[data_type])

    # Raw/opaque types: best-effort byte width, no decoding.
    return ScalarDefinition(size=resolved.get("size", 1))


def _is_structured_data_type(
    prop_data: dict[str, Any], definitions: dict[str, Any]
) -> bool:
    """Return True when a property's data spec contains an object or array.

    Used to decide whether to also build a PropertyValueDefinition tree for
    it: properties fully represented by a flat EntityDefinition (plain
    number/state/numericValue) do not need a duplicate structured
    representation.
    """
    data_spec = prop_data["data"]
    candidates = data_spec.get("oneOf", [data_spec])
    for candidate in candidates:
        resolved = _resolve_ref(candidate, definitions)
        if resolved.get("type") in ("object", "array"):
            return True
    return False


def _build_entities_from_object_property(
    class_code: int,
    epc: int,
    prop_data: dict[str, Any],
    definitions: dict[str, Any],
    atomic_paired_epcs: frozenset[int],
) -> list[EntityDefinition] | None:
    """Split a fixed-layout MRA 'object' property into flat EntityDefinitions.

    Each scalar numeric field becomes its own EntityDefinition with an
    auto-computed byte_offset, following the same id/key scheme as
    custom_definitions.yaml's manual byte_offset entries.

    Returns None when the property is not a fixed-layout object of scalar
    number fields (e.g. it contains a variable-length "array" field), or when
    it participates in an MRA "atomic" pairing (e.g. EPC 0xB2's channel range
    selector paired with EPC 0xB3's variable-length list) — the list side is
    unsupported, so its range-selector counterpart is left dropped too rather
    than exposing a config entity with no way to read the paired result.
    """
    if epc in atomic_paired_epcs:
        return None

    data_spec = _resolve_ref(prop_data["data"], definitions)
    if data_spec.get("type") != "object":
        return None

    access = prop_data["accessRule"]
    get_val = access["get"]
    set_val = access["set"]
    is_readable = get_val in ("required", "required_c", "required_o", "optional")
    is_writable = set_val in ("required", "required_c", "required_o", "optional")
    assert is_readable or is_writable, (
        f"Property for class 0x{class_code:04X} EPC 0x{epc:02X} "
        "is neither readable nor writable"
    )

    prop_name_en = _normalize_trailing_number(prop_data["propertyName"]["en"])
    prop_name_ja = prop_data["propertyName"]["ja"]

    entities: list[EntityDefinition] = []
    byte_offset = 0
    for field in data_spec["properties"]:
        element = _resolve_ref(field["element"], definitions)
        specs_to_check = element.get("oneOf", [element])
        resolved = [_resolve_ref(s, definitions) for s in specs_to_check]
        number_spec = next((s for s in resolved if s.get("type") == "number"), None)
        if number_spec is None:
            return None  # unsupported field shape (e.g. variable-length array)

        mra_format = number_spec.get("format")
        if not isinstance(mra_format, str) or mra_format not in _OBJECT_FIELD_BYTE_SIZE:
            return None
        byte_count = _OBJECT_FIELD_BYTE_SIZE[mra_format]

        coefficient_epcs = tuple(int(c, 16) for c in number_spec.get("coefficient", []))
        entity_id = (
            f"class_{class_code:04x}_epc_{epc:02x}"
            if byte_offset == 0
            else f"class_{class_code:04x}_epc_{epc:02x}_{byte_offset:02x}"
        )
        entities.append(
            EntityDefinition(
                id=entity_id,
                epc=epc,
                name_en=_normalize_trailing_number(
                    f"{prop_name_en} - {field['elementName']['en']}"
                ),
                name_ja=f"{prop_name_ja} {field['elementName']['ja']}",
                get=get_val,
                set=set_val,
                description_en=None,
                description_ja=None,
                format=mra_format,
                unit=number_spec.get("unit"),
                minimum=number_spec.get("minimum"),
                maximum=number_spec.get("maximum"),
                multiple_of=(
                    number_spec.get("multiple") or number_spec.get("multipleOf") or 1.0
                ),
                enum_values=(),
                byte_offset=byte_offset,
                coefficient_epcs=coefficient_epcs or None,
            )
        )
        byte_offset += byte_count

    return entities


# ============================================================================
# Definition Generation Functions
# ============================================================================


def _load_mra_metadata(mra_path: Path) -> tuple[str, dict[str, Any]]:
    """Load MRA metadata and definitions."""
    with (mra_path / "metaData.json").open(encoding="utf-8") as f:
        metadata = json.load(f)
    mra_version = metadata["metaData"]["dataVersion"]

    with (mra_path / "definitions" / "definitions.json").open(encoding="utf-8") as f:
        defs_data = json.load(f)
    mra_definitions = defs_data["definitions"]

    return mra_version, mra_definitions


def _is_latest_version(prop_data: dict[str, Any]) -> bool:
    """Check if property is valid for the latest MRA version.

    MRA properties have validRelease.to field indicating version validity.
    We only include properties valid for the latest version to avoid duplicates.
    """
    to_value: str = prop_data["validRelease"]["to"]
    return to_value == "latest"


# ============================================================================
# Property Roles Loading
#
# scripts/property_roles.xlsx is the curated source of truth for
# EntityDefinition.role, keyed by (class_code, epc) — class_code 0 denotes a
# common (superClass) property, matching the class_code already used for
# `common` entities above. It is maintained via scripts/sync_property_roles.py
# and manual review in a spreadsheet editor, not by this generator.
# ============================================================================

_ROLE_SHEET_NAME = "Roles"


def _load_property_roles(path: Path) -> dict[tuple[int, int], PropertyRole]:
    """Load curated (class_code, epc) -> PropertyRole from an xlsx file.

    Blank ``role`` cells are omitted from the result so that callers fall
    back to the ``EntityDefinition.role`` dataclass default (PRIMARY).
    Returns an empty dict (all entities default to PRIMARY) if the file does
    not exist yet.
    """
    if not path.exists():
        print(f"  Warning: {path} not found; all entities default to PRIMARY")
        return {}

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[_ROLE_SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    col = {name: idx for idx, name in enumerate(header)}

    roles: dict[tuple[int, int], PropertyRole] = {}
    for row in rows:
        class_code = _parse_hex_int(row[col["class_code"]])
        epc = _parse_hex_int(row[col["epc"]])
        role_text = row[col["role"]]
        if class_code is None or epc is None or not role_text:
            continue
        roles[class_code, epc] = PropertyRole(str(role_text).strip().lower())

    workbook.close()
    return roles


def _apply_role(
    entity: EntityDefinition,
    role_map: dict[tuple[int, int], PropertyRole],
    class_code: int,
) -> EntityDefinition:
    """Return ``entity`` with its curated role applied, if one is on file."""
    role = role_map.get((class_code, entity.epc))
    return entity if role is None else dataclasses.replace(entity, role=role)


def generate_definitions(mra_path: Path) -> _DefinitionsBuild:
    """Generate definitions from MRA data."""
    devices_path = mra_path / "devices"
    if not devices_path.exists():
        print(f"Error: devices directory not found at {devices_path}")
        return _DefinitionsBuild(
            version="1.0.0",
            mra_version="unknown",
            common=[],
            devices={},
            manufacturers={},
            mra_epcs={},
            common_epcs=frozenset(),
        )

    mra_version, mra_definitions = _load_mra_metadata(mra_path)
    role_map = _load_property_roles(PROPERTY_ROLES_FILE)
    print(f"  Loaded {len(role_map)} curated property role(s)")

    # Build common entities once (shared across all device classes)
    with (mra_path / "superClass" / "0x0000.json").open(encoding="utf-8") as f:
        superclass = json.load(f)
    common: list[EntityDefinition] = []
    raw_common_epcs: set[int] = set()
    for prop_data in superclass["elProperties"]:
        if not _is_latest_version(prop_data):
            continue
        raw_common_epcs.add(int(prop_data["epc"], 16))
        prop = _parse_mra_property(prop_data, mra_definitions)
        entity = _build_entity_from_property(0, prop)
        if entity is not None:
            common.append(_apply_role(entity, role_map, 0))

    # EPCs already in common section (to avoid duplicates in device-specific entities)
    common_entity_epcs = frozenset(e.epc for e in common)

    devices: dict[int, _DeviceBuild] = {}
    mra_epcs: dict[int, frozenset[int]] = {}

    for device_file in sorted(devices_path.glob("0x*.json")):
        class_code = int(device_file.stem, 16)

        with device_file.open(encoding="utf-8") as f:
            data = json.load(f)

        class_name_data = data["className"]
        mra_epcs[class_code] = frozenset(
            int(prop_data["epc"], 16)
            for prop_data in data["elProperties"]
            if _is_latest_version(prop_data)
        )
        entities: list[EntityDefinition] = []
        structured_values: dict[int, PropertyValueDefinition] = {}

        # EPCs participating in an MRA "atomic" pairing (e.g. a channel-range
        # selector paired with a variable-length list result). Both sides are
        # excluded from object splitting; see
        # _build_entities_from_object_property().
        atomic_paired_epcs: frozenset[int] = frozenset(
            epc
            for p in data["elProperties"]
            if "atomic" in p
            for epc in (int(p["epc"], 16), int(p["atomic"], 16))
        )

        for prop_data in data["elProperties"]:
            # Only include properties valid for the latest MRA version
            if not _is_latest_version(prop_data):
                continue

            epc = int(prop_data["epc"], 16)

            # Skip common EPCs (they are in the common section)
            if epc in common_entity_epcs:
                continue

            object_entities = _build_entities_from_object_property(
                class_code, epc, prop_data, mra_definitions, atomic_paired_epcs
            )
            if object_entities is not None:
                entities.extend(
                    _apply_role(e, role_map, class_code) for e in object_entities
                )
                continue

            prop = _parse_mra_property(prop_data, mra_definitions)
            entity = _build_entity_from_property(class_code, prop)
            if entity:
                entities.append(_apply_role(entity, role_map, class_code))
                continue

            # Neither a flat scalar entity nor a fixed-layout object split
            # could represent this property. Preserve a recursive value tree
            # for it when it genuinely contains an array/object structure
            # (e.g. a variable-length channel list), rather than dropping it
            # entirely.
            if _is_structured_data_type(prop_data, mra_definitions):
                structured_values[epc] = _parse_value_spec(
                    prop_data["data"], mra_definitions
                )

        # Register all MRA device classes, even those without device-specific
        # entities, so common entities (e.g., operation status 0x80) are
        # still applied via _load_devices().
        devices[class_code] = _DeviceBuild(
            name_en=class_name_data["en"],
            name_ja=class_name_data["ja"],
            entities=entities,
            structured_values=structured_values,
        )

    manufacturers = _load_manufacturer_codes(MANUFACTURER_CODES_FILE)

    return _DefinitionsBuild(
        version="1.0.0",
        mra_version=mra_version,
        common=common,
        devices=devices,
        manufacturers=manufacturers,
        mra_epcs=mra_epcs,
        common_epcs=frozenset(raw_common_epcs),
    )


# ============================================================================
# Manufacturer Codes Loading
# ============================================================================


def _load_manufacturer_codes(path: Path) -> dict[int, ManufacturerDefinition]:
    """Load manufacturer codes from YAML.

    Returns a mapping of manufacturer code (int) to ManufacturerDefinition.
    Returns an empty dict if the file does not exist.
    """
    if not path.exists():
        return {}

    with path.open(encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}

    result: dict[int, ManufacturerDefinition] = {}
    for entry in data.get("manufacturers", []):
        code = entry.get("manufacturer_code")
        if code is None:
            continue
        result[int(code)] = ManufacturerDefinition(
            name_en=entry.get("name_en") or "",
            name_ja=entry.get("name_ja") or "",
        )
    return result


# ============================================================================
# Custom Definitions Loading
# ============================================================================


def _load_custom_definitions(custom_path: Path) -> dict[str, Any]:
    """Load custom definitions from YAML file.

    Args:
        custom_path: Path to custom_definitions.yaml.

    Returns:
        Parsed custom definitions or empty dict if not found.
    """
    if not custom_path.exists():
        return {}

    with custom_path.open(encoding="utf-8") as f:
        data = yaml.safe_load(f)

    return data if data else {}


# YAML field name → EntityDefinition attribute name (where they differ)
_YAML_TO_ATTR: dict[str, str] = {"multipleOf": "multiple_of"}
_ACCESS_VALUES = frozenset(
    {"required", "required_c", "required_o", "optional", "notApplicable"}
)
_NUMERIC_FORMATS = frozenset({"uint8", "int8", "uint16", "int16", "uint32", "int32"})
_PATCH_FIELDS = frozenset(
    {
        "name_en",
        "name_ja",
        "get",
        "set",
        "description_en",
        "description_ja",
        "format",
        "unit",
        "minimum",
        "maximum",
        "multipleOf",
        "role",
        "enum_values",
    }
)
_DEFINE_FIELDS = _PATCH_FIELDS | frozenset(
    {"epc", "manufacturer_code", "byte_offset", "mode"}
)


def _custom_error(class_code: int, entry: dict[str, Any], message: str) -> ValueError:
    """Build a contextual custom definitions validation error."""
    epc = entry.get("epc")
    epc_text = f" EPC 0x{epc:02X}" if isinstance(epc, int) else ""
    return ValueError(
        f"Custom definition class 0x{class_code:04X}{epc_text}: {message}"
    )


def _validate_code(
    value: Any, maximum: int, field: str, class_code: int, entry: dict[str, Any]
) -> int:
    """Validate an integer code in a custom definition."""
    if (
        not isinstance(value, int)
        or isinstance(value, bool)
        or not 0 <= value <= maximum
    ):
        raise _custom_error(class_code, entry, f"{field} must be 0x0-{maximum:X}")
    return value


def _validate_access(
    value: Any, field: str, class_code: int, entry: dict[str, Any]
) -> str:
    """Validate an ECHONET Lite access rule."""
    if not isinstance(value, str) or value not in _ACCESS_VALUES:
        raise _custom_error(class_code, entry, f"{field} has an invalid access rule")
    return value


def _validate_numeric_fields(
    values: dict[str, Any], class_code: int, entry: dict[str, Any]
) -> None:
    """Validate numeric definition fields supplied by a custom entry."""
    if "format" in values and values["format"] not in _NUMERIC_FORMATS:
        raise _custom_error(class_code, entry, "format is not supported")
    for field in ("minimum", "maximum", "multipleOf"):
        if field in values and (
            not isinstance(values[field], int | float)
            or isinstance(values[field], bool)
        ):
            raise _custom_error(class_code, entry, f"{field} must be numeric")
    if (
        "minimum" in values
        and "maximum" in values
        and values["minimum"] > values["maximum"]
    ):
        raise _custom_error(class_code, entry, "minimum must not exceed maximum")
    if "multipleOf" in values and values["multipleOf"] <= 0:
        raise _custom_error(class_code, entry, "multipleOf must be greater than zero")


def _apply_enum_patch(
    entity: EntityDefinition, values: Any, class_code: int, entry: dict[str, Any]
) -> tuple[EnumValue, ...]:
    """Apply display-name patches to existing enum values."""
    if not isinstance(values, list):
        raise _custom_error(class_code, entry, "enum_values must be a list")

    by_key = {value.key: value for value in entity.enum_values}
    updated = dict(by_key)
    patch_order: list[str] = []
    for patch in values:
        if not isinstance(patch, dict) or set(patch) - {"key", "name_en", "name_ja"}:
            raise _custom_error(
                class_code,
                entry,
                "enum patch values may only contain key, name_en, and name_ja",
            )
        key = patch.get("key")
        if not isinstance(key, str) or key not in by_key:
            raise _custom_error(class_code, entry, f"enum key {key!r} does not exist")
        if key in patch_order:
            raise _custom_error(class_code, entry, f"duplicate enum key {key!r}")
        if "name_en" not in patch and "name_ja" not in patch:
            raise _custom_error(
                class_code, entry, "enum patch must change a display name"
            )
        current = by_key[key]
        updated[key] = dataclasses.replace(
            current,
            name_en=patch.get("name_en", current.name_en),
            name_ja=patch.get("name_ja", current.name_ja),
        )
        patch_order.append(key)
    if len(patch_order) == len(by_key):
        return tuple(updated[key] for key in patch_order)
    return tuple(updated[value.key] for value in entity.enum_values)


def _apply_patch(
    build: _DefinitionsBuild, class_code: int, entry: dict[str, Any]
) -> int:
    """Apply a partial custom patch to an MRA entity."""
    if "manufacturer_code" in entry:
        raise _custom_error(
            class_code, entry, "patch entries may not specify manufacturer_code"
        )
    epc = _validate_code(entry.get("epc"), 0xFF, "epc", class_code, entry)
    if epc not in build.mra_epcs.get(class_code, frozenset()):
        raise _custom_error(class_code, entry, "patch target is not an MRA EPC")
    if epc in build.common_epcs:
        raise _custom_error(class_code, entry, "common EPCs may not be patched")
    if unknown := set(entry) - (_PATCH_FIELDS | {"epc", "mode", "byte_offset"}):
        raise _custom_error(
            class_code, entry, f"unknown patch fields: {sorted(unknown)}"
        )
    changes_raw = {key: value for key, value in entry.items() if key in _PATCH_FIELDS}
    if not changes_raw:
        raise _custom_error(class_code, entry, "patch must change at least one field")
    _validate_numeric_fields(changes_raw, class_code, entry)

    device_build = build.devices[class_code]
    candidates = [
        (index, entity)
        for index, entity in enumerate(device_build.entities)
        if entity.epc == epc
    ]
    if not candidates:
        raise _custom_error(
            class_code, entry, "patch target is not a scalar MRA entity"
        )
    if len(candidates) > 1 and "byte_offset" not in entry:
        raise _custom_error(
            class_code,
            entry,
            "patch target has multiple byte offsets; specify byte_offset",
        )
    if "byte_offset" in entry:
        byte_offset = _validate_code(
            entry["byte_offset"], 0xFF, "byte_offset", class_code, entry
        )
        candidates = [
            (index, entity)
            for index, entity in candidates
            if entity.byte_offset == byte_offset
        ]
        if not candidates:
            raise _custom_error(class_code, entry, "patch byte_offset does not exist")

    for index, entity in candidates:
        numeric_fields = {"format", "unit", "minimum", "maximum", "multipleOf"}
        if entity.enum_values and numeric_fields & changes_raw.keys():
            raise _custom_error(
                class_code, entry, "numeric fields may not patch an enum entity"
            )
        if "enum_values" in changes_raw and not entity.enum_values:
            raise _custom_error(
                class_code, entry, "enum_values may only patch an enum entity"
            )
        changes: dict[str, Any] = {}
        for field, value in changes_raw.items():
            if field == "enum_values":
                changes[field] = _apply_enum_patch(entity, value, class_code, entry)
            elif field == "role":
                changes[field] = PropertyRole(value)
            elif field in {"get", "set"}:
                changes[field] = _validate_access(value, field, class_code, entry)
            else:
                changes[_YAML_TO_ATTR.get(field, field)] = value
        patched = dataclasses.replace(entity, **changes)
        if numeric_fields & changes_raw.keys():
            numeric_values: dict[str, Any] = {"format": patched.format}
            for field in ("minimum", "maximum"):
                if (value := getattr(patched, field)) is not None:
                    numeric_values[field] = value
            if patched.multiple_of != 1.0:
                numeric_values["multipleOf"] = patched.multiple_of
            _validate_numeric_fields(numeric_values, class_code, entry)
        device_build.entities[index] = patched
    return len(candidates)


def _build_custom_entity(class_code: int, entry: dict[str, Any]) -> EntityDefinition:
    """Build a fully-specified custom EntityDefinition from a define entry."""
    if unknown := set(entry) - _DEFINE_FIELDS:
        raise _custom_error(
            class_code, entry, f"unknown define fields: {sorted(unknown)}"
        )
    epc = _validate_code(entry.get("epc"), 0xFF, "epc", class_code, entry)
    mfr_code = (
        _validate_code(
            entry["manufacturer_code"], 0xFFFFFF, "manufacturer_code", class_code, entry
        )
        if "manufacturer_code" in entry
        else None
    )
    byte_offset = _validate_code(
        entry.get("byte_offset", 0), 0xFF, "byte_offset", class_code, entry
    )
    for field in ("name_en", "name_ja"):
        if not isinstance(entry.get(field), str) or not entry[field]:
            raise _custom_error(class_code, entry, f"{field} is required for define")
    get = _validate_access(entry.get("get"), "get", class_code, entry)
    set_value = _validate_access(
        entry.get("set", "notApplicable"), "set", class_code, entry
    )
    enum_values_raw = entry.get("enum_values")
    has_format = "format" in entry
    if bool(enum_values_raw) == has_format:
        raise _custom_error(
            class_code, entry, "define requires exactly one of format or enum_values"
        )
    if enum_values_raw is not None and not isinstance(enum_values_raw, list):
        raise _custom_error(class_code, entry, "enum_values must be a list")
    _validate_numeric_fields(entry, class_code, entry)

    enum_values: tuple[EnumValue, ...] = ()
    if enum_values_raw:
        enum_keys: set[str] = set()
        enum_edts: set[int] = set()
        enum_values = tuple(
            _build_enum_value(value, class_code, entry, enum_keys, enum_edts)
            for value in enum_values_raw
        )

    role = PropertyRole(entry["role"]) if "role" in entry else PropertyRole.PRIMARY
    entity_id = f"class_{class_code:04x}_epc_{epc:02x}"
    if mfr_code is not None:
        entity_id += f"_{mfr_code:06x}"
    if byte_offset:
        entity_id += f"_{byte_offset:02x}"
    return EntityDefinition(
        id=entity_id,
        epc=epc,
        name_en=entry["name_en"],
        name_ja=entry["name_ja"],
        get=get,
        set=set_value,
        description_en=entry.get("description_en"),
        description_ja=entry.get("description_ja"),
        format=entry.get("format") if not enum_values else None,
        unit=entry.get("unit") if not enum_values else None,
        minimum=entry.get("minimum") if not enum_values else None,
        maximum=entry.get("maximum") if not enum_values else None,
        multiple_of=entry.get("multipleOf", 1.0) if not enum_values else 1.0,
        enum_values=_normalize_two_value_enum_keys(
            enum_values, class_code, epc, mfr_code
        ),
        byte_offset=byte_offset,
        manufacturer_code=mfr_code,
        role=role,
    )


def _build_enum_value(
    value: Any,
    class_code: int,
    entry: dict[str, Any],
    enum_keys: set[str],
    enum_edts: set[int],
) -> EnumValue:
    """Validate and build a fully-specified enum value."""
    if not isinstance(value, dict) or set(value) != {
        "edt",
        "key",
        "name_en",
        "name_ja",
    }:
        raise _custom_error(
            class_code,
            entry,
            "define enum values require edt, key, name_en, and name_ja",
        )
    edt = _validate_code(value["edt"], 0xFF, "enum edt", class_code, entry)
    key = value["key"]
    if not isinstance(key, str) or not key:
        raise _custom_error(class_code, entry, "enum key must be a non-empty string")
    if key in enum_keys or edt in enum_edts:
        raise _custom_error(class_code, entry, "enum keys and EDTs must be unique")
    if not all(
        isinstance(value[name], str) and value[name] for name in ("name_en", "name_ja")
    ):
        raise _custom_error(
            class_code, entry, "enum display names must be non-empty strings"
        )
    enum_keys.add(key)
    enum_edts.add(edt)
    return EnumValue(
        edt=edt, key=key, name_en=value["name_en"], name_ja=value["name_ja"]
    )


def _apply_custom_definitions(build: _DefinitionsBuild, custom: dict[str, Any]) -> None:
    """Apply unified custom class definitions to an MRA definitions build."""
    devices = custom.get("devices", {})
    if not isinstance(devices, dict):
        raise ValueError("Custom definitions devices must be a mapping")

    define_count = 0
    patch_count = 0
    for class_code_raw, class_entry in devices.items():
        class_code = _validate_code(
            class_code_raw, 0xFFFF, "class_code", 0, {"epc": class_code_raw}
        )
        if not isinstance(class_entry, dict):
            raise ValueError(
                f"Custom definition class 0x{class_code:04X} must be a mapping"
            )
        if unknown := set(class_entry) - {"name_en", "name_ja", "properties"}:
            raise ValueError(
                f"Custom definition class 0x{class_code:04X} has unknown fields: "
                f"{sorted(unknown)}"
            )
        properties = class_entry.get("properties")
        if not isinstance(properties, list) or not properties:
            raise ValueError(
                f"Custom definition class 0x{class_code:04X} requires properties"
            )

        if class_code not in build.devices:
            names = (class_entry.get("name_en"), class_entry.get("name_ja"))
            if not all(isinstance(name, str) and name for name in names):
                raise ValueError(
                    f"Custom definition class 0x{class_code:04X} requires name_en and name_ja"
                )
            build.devices[class_code] = _DeviceBuild(
                name_en=class_entry["name_en"],
                name_ja=class_entry["name_ja"],
                entities=[],
            )
            build.mra_epcs[class_code] = frozenset()
        elif "name_en" in class_entry or "name_ja" in class_entry:
            if not all(
                isinstance(class_entry.get(field), str) and class_entry[field]
                for field in ("name_en", "name_ja")
            ):
                raise ValueError(
                    f"Custom definition class 0x{class_code:04X} requires both names"
                )
            device_build = build.devices[class_code]
            device_build.name_en = class_entry["name_en"]
            device_build.name_ja = class_entry["name_ja"]

        define_keys: set[tuple[int, int, int | None]] = set()
        enum_keys: set[tuple[int, int | None]] = set()
        for entry in properties:
            if not isinstance(entry, dict):
                raise ValueError(
                    f"Custom definition class 0x{class_code:04X} properties must be mappings"
                )
            mode = entry.get("mode", "define")
            if mode == "patch":
                patch_count += _apply_patch(build, class_code, entry)
                continue
            if mode != "define":
                raise _custom_error(class_code, entry, "mode must be define or patch")

            epc = _validate_code(entry.get("epc"), 0xFF, "epc", class_code, entry)
            if epc in build.mra_epcs.get(class_code, frozenset()):
                raise _custom_error(
                    class_code, entry, "define target is already an MRA EPC"
                )
            byte_offset = _validate_code(
                entry.get("byte_offset", 0), 0xFF, "byte_offset", class_code, entry
            )
            manufacturer_code = entry.get("manufacturer_code")
            if manufacturer_code is not None:
                manufacturer_code = _validate_code(
                    manufacturer_code, 0xFFFFFF, "manufacturer_code", class_code, entry
                )
            define_key = (epc, byte_offset, manufacturer_code)
            if define_key in define_keys:
                raise _custom_error(class_code, entry, "duplicate define")
            define_keys.add(define_key)
            matching_scopes = {
                scope
                for defined_epc, defined_offset, scope in define_keys
                if defined_epc == epc and defined_offset == byte_offset
            }
            if None in matching_scopes and len(matching_scopes) > 1:
                raise _custom_error(
                    class_code,
                    entry,
                    "generic and manufacturer-specific defines may not coexist",
                )
            if "enum_values" in entry:
                enum_key = (epc, manufacturer_code)
                if enum_key in enum_keys:
                    raise _custom_error(
                        class_code,
                        entry,
                        "multiple enum defines for an EPC and manufacturer are not supported",
                    )
                enum_keys.add(enum_key)
            build.devices[class_code].entities.append(
                _build_custom_entity(class_code, entry)
            )
            define_count += 1

    if patch_count:
        print(f"  Applied {patch_count} patch(es)")
    if define_count:
        print(f"  Added {define_count} custom definition(s)")


def _load_collection_bindings(
    custom: dict[str, Any],
) -> dict[int, list[CollectionBinding]]:
    """Load curated CollectionBinding entries from custom_definitions.yaml.

    MRA does not document which count property backs which paged list
    result (e.g. that class 0x0287 EPC 0xBE is paged using EPC 0xB8's
    channel count), so this relationship is curated by hand under the
    top-level ``collection_bindings`` key.
    """
    bindings: dict[int, list[CollectionBinding]] = {}
    for entry in custom.get("collection_bindings", []):
        class_code: int = entry["class_code"]
        bindings.setdefault(class_code, []).append(
            CollectionBinding(
                result_epc=entry["result_epc"],
                count_epc=entry.get("count_epc"),
                items_path=tuple(entry["items_path"]),
                start_path=tuple(entry["start_path"]),
                page_count_path=tuple(entry["page_count_path"]),
                index_kind=CollectionIndex(entry.get("index_kind", "channel")),
            )
        )
    return bindings


# ============================================================================
# Code Generation Functions
# ============================================================================


def _validate_entity(entity: EntityDefinition, class_code: int) -> None:
    """Validate an entity definition at build time."""
    assert entity.enum_values or entity.format or entity.numeric_values, (
        f"Entity EPC 0x{entity.epc:02X} for class 0x{class_code:04X} missing format"
    )
    assert (
        not entity.enum_values
        or len(entity.enum_values) != 1
        or entity.get == "notApplicable"
        or entity.format is not None
    ), (
        f"Entity EPC 0x{entity.epc:02X} for class 0x{class_code:04X}"
        " has only 1 enum_value"
    )


def _generate_python_source(build: _DefinitionsBuild) -> str:
    """Render definitions as importable Python source.

    Produces a module that builds the DefinitionsRegistry from dataclass
    literals (rendered via repr), so the data is referenced directly as code
    instead of being parsed from JSON at runtime. Common entities are emitted
    once as ``_COMMON`` and shared by every device class.
    """
    # Build-time validation
    for class_code, device_build in build.devices.items():
        for entity in build.common + device_build.entities:
            _validate_entity(entity, class_code)

    lines: list[str] = [
        "# ruff: noqa",
        '"""Auto-generated ECHONET Lite definitions.',
        "",
        "DO NOT EDIT. Generated by scripts/generate_definitions.py.",
        '"""',
        "",
        "from .definitions import (",
        "    ArrayDefinition,",
        "    CollectionBinding,",
        "    CollectionIndex,",
        "    DefinitionsRegistry,",
        "    DeviceDefinition,",
        "    EntityDefinition,",
        "    EnumValue,",
        "    ManufacturerDefinition,",
        "    NumericValueEntry,",
        "    ObjectDefinition,",
        "    ObjectField,",
        "    OneOfDefinition,",
        "    PropertyRole,",
        "    PropertyValueDefinition,",
        "    ScalarDefinition,",
        ")",
        "",
        "_COMMON: tuple[EntityDefinition, ...] = (",
    ]
    lines.extend(f"    {entity!r}," for entity in build.common)
    lines.extend(
        [
            ")",
            "",
            "def _merge_entities(",
            "    *groups: tuple[EntityDefinition, ...],",
            ") -> tuple[EntityDefinition, ...]:",
            '    """Merge entity groups, letting later groups replace matching EPCs."""',
            "    merged: list[EntityDefinition] = []",
            "    for group in groups:",
            "        epcs = {entity.epc for entity in group}",
            "        merged = [entity for entity in merged if entity.epc not in epcs]",
            "        merged.extend(group)",
            "    return tuple(merged)",
            "",
        ]
    )
    lines.append("DEVICES: dict[int, DeviceDefinition] = {")
    for class_code in sorted(build.devices):
        device_build = build.devices[class_code]
        lines.append(f"    {class_code}: DeviceDefinition(")
        lines.append(f"        class_code={class_code},")
        lines.append(f"        name_en={device_build.name_en!r},")
        lines.append(f"        name_ja={device_build.name_ja!r},")
        if device_build.entities:
            lines.append("        entities=_merge_entities(")
            lines.append("            _COMMON,")
            lines.append("            (")
            lines.extend(
                f"                {entity!r}," for entity in device_build.entities
            )
            lines.append("            ),")
            lines.append("        ),")
        else:
            lines.append("        entities=_merge_entities(_COMMON),")
        lines.append("    ),")
    lines.append("}")
    lines.append("")
    lines.append("MANUFACTURERS: dict[int, ManufacturerDefinition] = {")
    lines.extend(
        f"    {code}: {build.manufacturers[code]!r},"
        for code in sorted(build.manufacturers)
    )
    lines.append("}")
    lines.append("")
    lines.append("STRUCTURED_VALUES: dict[int, dict[int, PropertyValueDefinition]] = {")
    for class_code in sorted(build.devices):
        structured_values = build.devices[class_code].structured_values
        if not structured_values:
            continue
        lines.append(f"    {class_code}: {{")
        for epc in sorted(structured_values):
            lines.append(f"        {epc}: {structured_values[epc]!r},")
        lines.append("    },")
    lines.append("}")
    lines.append("")
    lines.append("COLLECTION_BINDINGS: dict[int, tuple[CollectionBinding, ...]] = {")
    for class_code in sorted(build.devices):
        collection_bindings = build.devices[class_code].collection_bindings
        if not collection_bindings:
            continue
        lines.append(f"    {class_code}: (")
        lines.extend(f"        {binding!r}," for binding in collection_bindings)
        lines.append("    ),")
    lines.append("}")
    lines.append("")
    lines.extend(
        [
            "REGISTRY = DefinitionsRegistry(",
            f"    version={build.version!r},",
            f"    mra_version={build.mra_version!r},",
            "    devices=DEVICES,",
            "    entities={cc: d.entities for cc, d in DEVICES.items()},",
            "    manufacturers=MANUFACTURERS,",
            "    structured_values=STRUCTURED_VALUES,",
            "    collection_bindings=COLLECTION_BINDINGS,",
            ")",
            "",
        ]
    )
    return "\n".join(lines)


# ============================================================================
# Main Entry Point
# ============================================================================


def main() -> None:
    """Main entry point."""
    if not MRA_DIR.exists():
        print(f"Error: MRA directory not found at {MRA_DIR}")
        print("Please ensure the mra/ directory exists with MRA data.")
        return

    print(f"Using MRA data from: {MRA_DIR}")
    print("Generating definitions...")
    build = generate_definitions(MRA_DIR)

    # Load and merge custom vendor definitions
    if CUSTOM_DEFINITIONS_FILE.exists():
        print(f"\nLoading custom definitions from {CUSTOM_DEFINITIONS_FILE}...")
        custom = _load_custom_definitions(CUSTOM_DEFINITIONS_FILE)
        _apply_custom_definitions(build, custom)

        for class_code, bindings in _load_collection_bindings(custom).items():
            if class_code not in build.devices:
                print(
                    f"  Warning: collection_bindings target class 0x{class_code:04X} not found"
                )
                continue
            build.devices[class_code].collection_bindings.extend(bindings)

    generated_source = _generate_python_source(build)
    generated_path = PYHEMS_DIR / "_definitions_generated.py"
    with generated_path.open("w", encoding="utf-8") as f:
        f.write(generated_source)
    print(f"\nGenerated: {generated_path}")

    device_count = len(build.devices)
    entity_count = sum(len(db.entities) for db in build.devices.values())
    manufacturer_count = len(build.manufacturers)
    print("\nSummary:")
    print(f"  MRA version: {build.mra_version}")
    print(f"  Devices: {device_count}")
    print(f"  Entities: {entity_count}")
    print(f"  Manufacturers: {manufacturer_count}")


if __name__ == "__main__":
    main()
