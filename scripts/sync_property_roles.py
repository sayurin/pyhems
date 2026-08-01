#!/usr/bin/env -S uv run
# /// script
# requires-python = ">=3.13"
# dependencies = [
#   "openpyxl",
#   "pyyaml",
# ]
# ///
# ruff: noqa: T201
"""Reconcile scripts/property_roles.xlsx against current MRA data.

This tool is independent from generate_definitions.py on purpose: after MRA
data changes, newly introduced (class_code, epc) pairs need a human to look
at them in a spreadsheet editor before generate_definitions.py can pick up
their role. Folding both steps into one script would either skip that human
step or block the build on it; neither is desirable.

Workflow:
    1. Update mra/ (MRA data refresh).
    2. Run this tool: uv run scripts/sync_property_roles.py
       - Adds rows for new (class_code, epc) pairs, pre-filled with a
         rule-based ``candidate_role`` suggestion (informational only).
       - Refreshes MRA-derived columns (names, description, get/set, unit,
         format) for existing rows without touching ``role``/``comment``.
       - Drops rows for properties no longer present in MRA.
    3. Open scripts/property_roles.xlsx in a spreadsheet editor, fill in
       ``role`` for rows whose status is NEW/PENDING/CHANGED.
    4. Run: uv run scripts/generate_definitions.py

Properties with a blank ``role`` default to PRIMARY (see
pyhems.definitions.PropertyRole) — this tool never blocks the build.

The ``status`` column is a spreadsheet formula, not a value written by this
script: it flips to OK the instant a reviewer fills in ``role``, without
re-running this tool. This tool only maintains two hidden columns
(``_is_new``, ``_mra_changed``) that the formula reads, since detecting
"did MRA change since the last run" requires comparing against MRA data
external to the spreadsheet and can only happen here.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

SCRIPTS_DIR = Path(__file__).parent
sys.path.insert(0, str(SCRIPTS_DIR))
import generate_definitions as gd  # noqa: E402

sys.path.pop(0)

MRA_DIR = gd.MRA_DIR
PROPERTY_ROLES_FILE = gd.PROPERTY_ROLES_FILE
ROLES_SHEET_NAME = "Roles"
INSTRUCTIONS_SHEET_NAME = "Instructions"

# Visible columns. All but ``role``/``comment`` (human-edited, preserved
# across runs) and ``status`` (a formula, see below) are refreshed from MRA
# on every run.
HEADER = [
    "class_code",
    "epc",
    "class_name_ja",
    "name_en",
    "name_ja",
    "description_ja",
    "get",
    "set",
    "unit",
    "format",
    "status",
    "candidate_role",
    "role",
    "comment",
]
# Hidden columns feeding the ``status`` formula only; never read by
# generate_definitions.py.
_IS_NEW_COL = "_is_new"
_MRA_CHANGED_COL = "_mra_changed"
HIDDEN_HEADER = [_IS_NEW_COL, _MRA_CHANGED_COL]
ALL_COLUMNS = HEADER + HIDDEN_HEADER

# Fields that trigger _mra_changed when they drift from the last recorded
# snapshot. Deliberately excludes class_name_ja/description_ja (cosmetic;
# doesn't affect the role decision).
_TRACKED_FIELDS = ("name_en", "get", "set", "unit", "format")

_ROLE_VALUES = ("primary", "instantaneous", "setting", "status", "specification")

# Keyword-based PropertyRole suggestions. These are informational hints for
# the reviewer only — generate_definitions.py never reads ``candidate_role``.
_SPECIFICATION_KEYWORDS = (
    "rated",
    "capacity",
    "number of effective digits",
    "tolerance class",
)
_STATUS_KEYWORDS = (
    "fault",
    "abnormal",
    "emergency",
    "exceptional",
    "maintenance",
    "filter change",
    "battery level",
    "remote control setting status",
)
_SETTING_KEYWORDS = (
    "reset",
    "threshold",
    "reservation",
    "timer",
    "mute",
    "volume",
    "limit",
)


def _candidate_role(name_en: str, set_val: str, unit: str | None) -> str:
    """Suggest a PropertyRole from name/access-rule keywords, or "" if none."""
    name = name_en.lower()
    writable = set_val != "notApplicable"
    readable_only = not writable
    if readable_only and any(k in name for k in _SPECIFICATION_KEYWORDS):
        return "specification"
    if readable_only and any(k in name for k in _STATUS_KEYWORDS):
        return "status"
    if readable_only and unit and "instantaneous" in name:
        return "instantaneous"
    if writable and any(k in name for k in _SETTING_KEYWORDS):
        return "setting"
    return ""


def _scan_mra_properties(mra_path: Path) -> dict[tuple[int, int], dict[str, Any]]:
    """Scan MRA data for the current (class_code, epc) universe.

    Mirrors generate_definitions.generate_definitions()'s notion of "common"
    vs. device-specific properties: common (superClass) properties are keyed
    under class_code 0 and skipped when encountered again per-device.
    """
    _, mra_definitions = gd._load_mra_metadata(mra_path)

    def _extract(prop_data: dict[str, Any], class_name_ja: str) -> dict[str, Any]:
        name_en = gd._normalize_trailing_number(prop_data["propertyName"]["en"])
        name_ja = prop_data["propertyName"]["ja"]
        access = prop_data["accessRule"]
        data_spec = prop_data["data"]
        if "oneOf" in data_spec:
            data_spec = data_spec["oneOf"][0]
        resolved = gd._resolve_ref(data_spec, mra_definitions)
        unit = resolved.get("unit") if resolved.get("type") == "number" else None
        mra_format = (
            resolved.get("format") if resolved.get("type") == "number" else None
        )
        return {
            "class_name_ja": class_name_ja,
            "name_en": name_en,
            "name_ja": name_ja,
            "description_ja": prop_data["descriptions"]["ja"],
            "get": access["get"],
            "set": access["set"],
            "unit": unit or "",
            "format": mra_format or "",
        }

    properties: dict[tuple[int, int], dict[str, Any]] = {}

    with (mra_path / "superClass" / "0x0000.json").open(encoding="utf-8") as f:
        superclass = json.load(f)
    superclass_name_ja = superclass["className"]["ja"]
    for prop_data in superclass["elProperties"]:
        if not gd._is_latest_version(prop_data):
            continue
        epc = int(prop_data["epc"], 16)
        properties[0, epc] = _extract(prop_data, superclass_name_ja)

    common_epcs = frozenset(epc for (cc, epc) in properties if cc == 0)

    for device_file in sorted((mra_path / "devices").glob("0x*.json")):
        class_code = int(device_file.stem, 16)
        with device_file.open(encoding="utf-8") as f:
            data = json.load(f)
        device_name_ja = data["className"]["ja"]
        for prop_data in data["elProperties"]:
            if not gd._is_latest_version(prop_data):
                continue
            epc = int(prop_data["epc"], 16)
            if epc in common_epcs:
                continue
            properties[class_code, epc] = _extract(prop_data, device_name_ja)

    return properties


def _load_existing_rows(path: Path) -> dict[tuple[int, int], dict[str, str]]:
    """Load existing "Roles" sheet rows, keyed by (class_code, epc).

    Column lookup is name-based and tolerates a sheet missing columns this
    version of the script expects (e.g. before class_name_ja/description_ja
    were introduced), so role/comment survive a column-layout migration.
    """
    if not path.exists():
        return {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    if ROLES_SHEET_NAME not in workbook.sheetnames:
        return {}
    sheet = workbook[ROLES_SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    col = {name: idx for idx, name in enumerate(header)}
    if "class_code" not in col or "epc" not in col:
        workbook.close()
        return {}

    existing: dict[tuple[int, int], dict[str, str]] = {}
    for row in rows:
        class_code = gd._parse_hex_int(row[col["class_code"]])
        epc = gd._parse_hex_int(row[col["epc"]])
        if class_code is None or epc is None:
            continue
        existing[class_code, epc] = {
            name: ("" if row[idx] is None else str(row[idx]))
            for name, idx in col.items()
        }
    workbook.close()
    return existing


def _header_matches(path: Path) -> bool:
    """Return True if ``path``'s "Roles" sheet header already matches ALL_COLUMNS."""
    if not path.exists():
        return False
    workbook = load_workbook(path, read_only=True, data_only=True)
    if ROLES_SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        return False
    sheet = workbook[ROLES_SHEET_NAME]
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    workbook.close()
    return list(header) == ALL_COLUMNS


def _status_formula(row_num: int) -> str:
    """Build the ``status`` formula for spreadsheet row ``row_num``.

    Reads only the ``role`` cell and the two hidden helper columns on the
    same row, so it recomputes live in the spreadsheet the instant a
    reviewer fills in ``role`` — no re-run of this tool required.
    """
    role_col = get_column_letter(HEADER.index("role") + 1)
    is_new_col = get_column_letter(len(HEADER) + HIDDEN_HEADER.index(_IS_NEW_COL) + 1)
    mra_changed_col = get_column_letter(
        len(HEADER) + HIDDEN_HEADER.index(_MRA_CHANGED_COL) + 1
    )
    return (
        f'=IF(TRIM({role_col}{row_num})="",'
        f'IF({is_new_col}{row_num},"NEW","PENDING"),'
        f'IF({mra_changed_col}{row_num},"CHANGED","OK"))'
    )


def _build_rows(
    current: dict[tuple[int, int], dict[str, Any]],
    existing: dict[tuple[int, int], dict[str, str]],
) -> tuple[dict[tuple[int, int], dict[str, Any]], int]:
    """Compute the full row dataset (all ALL_COLUMNS values except status)."""
    rows: dict[tuple[int, int], dict[str, Any]] = {}
    for key in sorted(current):
        class_code, epc = key
        props = current[key]
        prior = existing.get(key)
        candidate = _candidate_role(props["name_en"], props["set"], props["unit"])

        is_new = prior is None
        if prior is None:
            mra_changed = False
            role, comment = "", ""
        else:
            mra_changed = any(
                prior.get(field, "") != props[field] for field in _TRACKED_FIELDS
            )
            role, comment = prior.get("role", ""), prior.get("comment", "")

        rows[key] = {
            "class_code": f"0x{class_code:04X}",
            "epc": f"0x{epc:02X}",
            "class_name_ja": props["class_name_ja"],
            "name_en": props["name_en"],
            "name_ja": props["name_ja"],
            "description_ja": props["description_ja"],
            "get": props["get"],
            "set": props["set"],
            "unit": props["unit"],
            "format": props["format"],
            "candidate_role": candidate,
            "role": role,
            "comment": comment,
            _IS_NEW_COL: is_new,
            _MRA_CHANGED_COL: mra_changed,
        }

    removed_count = len(existing.keys() - current.keys())
    return rows, removed_count


def _write_instructions_sheet(sheet: Worksheet) -> None:
    """Write the static "Instructions" sheet content."""
    rows = [
        ("item", "description"),
        ("purpose", "Curate PropertyRole per (class_code, epc) for pyhems."),
        (
            "role values",
            "primary (default) / instantaneous (fast-poll candidate) / "
            "setting / status / specification. Pick from the dropdown.",
        ),
        (
            "status column",
            "Formula, not editable. NEW: newly introduced by the latest MRA "
            "sync, role still blank. PENDING: role still blank (not new). "
            "CHANGED: role is set but MRA data changed since — re-check. "
            "OK: role is set and MRA data unchanged. Filling in role flips "
            "NEW/PENDING to OK immediately, no re-run needed.",
        ),
        (
            "candidate_role column",
            "Rule-based suggestion only; sync_property_roles.py never "
            "writes this into 'role'. Confirm or override manually.",
        ),
        (
            "role column",
            "Human decision. Blank defaults to PRIMARY at generation time; "
            "generate_definitions.py never fails on a blank role.",
        ),
        ("comment column", "Free text: rationale, open questions, etc."),
    ]
    for row in rows:
        sheet.append(row)


def _apply_role_validation(sheet: Worksheet, last_row: int) -> None:
    """(Re)apply the role-column dropdown over the current data range."""
    role_col = get_column_letter(HEADER.index("role") + 1)
    sheet.data_validations.dataValidation = [
        dv
        for dv in sheet.data_validations.dataValidation
        if not str(dv.sqref).startswith(f"{role_col}")
    ]
    if last_row < 2:
        return
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(_ROLE_VALUES)}"',
        allow_blank=True,
    )
    dv.add(f"{role_col}2:{role_col}{last_row}")
    sheet.add_data_validation(dv)


def _rebuild_workbook(path: Path, rows: dict[tuple[int, int], dict[str, Any]]) -> None:
    """Write a fresh "Roles" sheet from scratch (column layout changed)."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = ROLES_SHEET_NAME
    sheet.append(ALL_COLUMNS)

    data_columns = [name for name in HEADER if name != "status"] + HIDDEN_HEADER
    col_idx = {name: ALL_COLUMNS.index(name) + 1 for name in ALL_COLUMNS}

    for row_num, key in enumerate(sorted(rows), start=2):
        row = rows[key]
        for name in data_columns:
            sheet.cell(row_num, col_idx[name]).value = row[name]
        sheet.cell(row_num, col_idx["status"]).value = _status_formula(row_num)

    for name in HIDDEN_HEADER:
        col_letter = get_column_letter(ALL_COLUMNS.index(name) + 1)
        sheet.column_dimensions[col_letter].hidden = True

    _apply_role_validation(sheet, sheet.max_row)

    instructions_sheet = workbook.create_sheet(INSTRUCTIONS_SHEET_NAME)
    _write_instructions_sheet(instructions_sheet)

    workbook.save(path)


def _update_workbook_in_place(
    path: Path, rows: dict[tuple[int, int], dict[str, Any]]
) -> None:
    """Update an existing "Roles" sheet's values without touching formatting."""
    workbook = load_workbook(path)
    sheet = workbook[ROLES_SHEET_NAME]

    col_idx = {name: HEADER.index(name) + 1 for name in HEADER}
    col_idx.update(
        {name: len(HEADER) + HIDDEN_HEADER.index(name) + 1 for name in HIDDEN_HEADER}
    )

    key_to_row: dict[tuple[int, int], int] = {}
    for row_num in range(2, sheet.max_row + 1):
        class_code = gd._parse_hex_int(sheet.cell(row_num, col_idx["class_code"]).value)
        epc = gd._parse_hex_int(sheet.cell(row_num, col_idx["epc"]).value)
        if class_code is not None and epc is not None:
            key_to_row[class_code, epc] = row_num

    # Delete rows for properties no longer in MRA, bottom-to-top so earlier
    # row numbers stay valid.
    for key, row_num in sorted(key_to_row.items(), key=lambda kv: -kv[1]):
        if key not in rows:
            sheet.delete_rows(row_num, 1)
    key_to_row = {key: n for key, n in key_to_row.items() if key in rows}

    refresh_columns = [c for c in HEADER if c not in ("status", "role", "comment")]
    for key, row_num in key_to_row.items():
        row = rows[key]
        for name in refresh_columns:
            sheet.cell(row_num, col_idx[name]).value = row[name]
        for name in HIDDEN_HEADER:
            sheet.cell(row_num, col_idx[name]).value = row[name]

    # Append brand-new rows at the bottom, preserving existing rows' order.
    next_row = sheet.max_row + 1
    for key in sorted(rows.keys() - key_to_row.keys()):
        row = rows[key]
        for name in (*refresh_columns, "role", "comment"):
            sheet.cell(next_row, col_idx[name]).value = row[name]
        for name in HIDDEN_HEADER:
            sheet.cell(next_row, col_idx[name]).value = row[name]
        key_to_row[key] = next_row
        next_row += 1

    for row_num in key_to_row.values():
        sheet.cell(row_num, col_idx["status"]).value = _status_formula(row_num)

    _apply_role_validation(sheet, sheet.max_row)

    workbook.save(path)


def sync_property_roles(mra_path: Path, roles_path: Path) -> None:
    """Reconcile ``roles_path`` against the current MRA property universe."""
    current = _scan_mra_properties(mra_path)
    existing = _load_existing_rows(roles_path)
    rows, removed_count = _build_rows(current, existing)

    if _header_matches(roles_path):
        _update_workbook_in_place(roles_path, rows)
    else:
        _rebuild_workbook(roles_path, rows)

    new_count = sum(1 for r in rows.values() if r[_IS_NEW_COL])
    changed_count = sum(
        1 for r in rows.values() if not r[_IS_NEW_COL] and r[_MRA_CHANGED_COL]
    )
    ok_or_pending = len(rows) - new_count - changed_count

    print(f"Reconciled {roles_path}")
    print(
        f"  NEW={new_count} CHANGED={changed_count} REMOVED={removed_count} "
        f"OK/PENDING={ok_or_pending} TOTAL={len(rows)}"
    )
    if new_count or changed_count:
        print("  Review NEW/CHANGED rows before running generate_definitions.py")


def main() -> None:
    """Main entry point."""
    if not MRA_DIR.exists():
        print(f"Error: MRA directory not found at {MRA_DIR}")
        return
    sync_property_roles(MRA_DIR, PROPERTY_ROLES_FILE)


if __name__ == "__main__":
    main()
